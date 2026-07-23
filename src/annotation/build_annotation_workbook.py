"""Stage 5 (pilot): build a human-annotation workbook for a small sample of generated notes.

Uses the v0 error taxonomy sketched in the project brief (fabrication / omission /
contradiction) as a placeholder -- Stage 4 will formalize this (severity levels, sub-types,
edge cases). Labels made against v0 may need reconciling once that lands.

Sample: 3 encounters (D2N068, D2N069, D2N070) x 3 strategies = 9 generated notes, all from a
single model version (gemini-3.6-flash) to avoid the mixed-model confound documented in
data/generated/GENERATION_NOTES.md.

Sheets:
  Instructions          -- taxonomy definitions and how to fill each sheet
  Notes_Reference        -- full generated note text, for context before sentence-level review
  Transcripts_Reference   -- full transcript + gold note per encounter, ground truth to check against
  Sentence_Annotations    -- one row per generated-note sentence; label fabrication/contradiction/none
  Omission_Check          -- one row per (encounter, strategy); free-text list of missing facts
  Progress                -- live formula-based counts of what's been labeled so far
"""
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[2]
PROCESSED_DIR = ROOT / "data" / "processed"
GENERATED_DIR = ROOT / "data" / "generated"
OUT_PATH = ROOT / "data" / "annotation" / "stage5_pilot_annotation.xlsx"

SAMPLE_ENCOUNTERS = ["D2N068", "D2N069", "D2N070"]
STRATEGIES = ["zero_shot", "schema_guided", "few_shot"]
SPLIT = "valid"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def load_processed(split: str) -> dict:
    path = PROCESSED_DIR / f"{split}.jsonl"
    with open(path, encoding="utf-8") as f:
        return {json.loads(l)["encounter_id"]: json.loads(l) for l in f}


def load_generated(strategy: str, split: str, provider: str = "gemini") -> dict:
    path = GENERATED_DIR / f"{provider}_{strategy}_{split}.jsonl"
    with open(path, encoding="utf-8") as f:
        return {json.loads(l)["encounter_id"]: json.loads(l) for l in f}


def format_dialogue(turns: list[dict]) -> str:
    return "\n".join(f"[{t['speaker']}] {t['text']}" for t in turns)


SENTENCE_SPLIT_RE = re.compile(r"(?<=[.!?])\s+")


def split_note_into_units(note_text: str) -> list[tuple[str, str]]:
    """Return (kind, text) pairs: kind is 'header' or 'sentence'."""
    units = []
    for line in note_text.split("\n"):
        line = line.strip()
        if not line:
            continue
        if line.isupper() and len(line) < 60:
            units.append(("header", line))
            continue
        for sentence in SENTENCE_SPLIT_RE.split(line):
            sentence = sentence.strip()
            if sentence:
                units.append(("sentence", sentence))
    return units


def style_header_row(ws, row: int, ncols: int):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name=FONT_NAME, bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def build():
    processed = load_processed(SPLIT)
    generated = {s: load_generated(s, SPLIT) for s in STRATEGIES}

    wb = Workbook()

    # --- Instructions ---
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False
    lines = [
        ("Stage 5 pilot annotation -- v0 taxonomy", True),
        ("", False),
        ("This uses a placeholder taxonomy from the project brief. Stage 4 will formalize "
         "it (severity levels, sub-types); some labels here may need revisiting after that.", False),
        ("", False),
        ("FABRICATION", True),
        ("A statement in the note that was never said or implied anywhere in the transcript.", False),
        ("", False),
        ("CONTRADICTION", True),
        ("A statement in the note that states the opposite of, or is materially inconsistent "
         "with, something said in the transcript (e.g. wrong dosage, wrong symptom polarity).", False),
        ("", False),
        ("OMISSION", True),
        ("A clinically relevant fact stated in the transcript that is missing from the note. "
         "Checked per whole note in the Omission_Check sheet, not sentence-by-sentence, since "
         "an omission has no corresponding sentence in the note to attach to.", False),
        ("", False),
        ("HOW TO USE THIS WORKBOOK", True),
        ("1. Read Notes_Reference and Transcripts_Reference for the encounter/strategy you're "
         "labeling, so you have full context before judging individual sentences.", False),
        ("2. In Sentence_Annotations, go row by row through each generated-note sentence. Use "
         "the error_type dropdown (leave blank / choose 'none' if the sentence is faithful). "
         "If fabrication or contradiction, quote the relevant transcript excerpt (or note its "
         "absence, for fabrication) in evidence_quote.", False),
        ("3. In Omission_Check, for each (encounter, strategy) pair, list any clinically "
         "relevant facts from the transcript that the note left out.", False),
        ("4. Check Progress for a live count of what's labeled so far.", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(name=FONT_NAME, bold=bold, size=13 if bold and i == 1 else 11)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110

    # --- Notes_Reference ---
    ws = wb.create_sheet("Notes_Reference")
    headers = ["encounter_id", "strategy", "generated_note_full_text"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    row = 2
    for enc_id in SAMPLE_ENCOUNTERS:
        for strategy in STRATEGIES:
            note = generated[strategy][enc_id]["generated_note"]
            ws.cell(row=row, column=1, value=enc_id).font = Font(name=FONT_NAME)
            ws.cell(row=row, column=2, value=strategy).font = Font(name=FONT_NAME)
            c = ws.cell(row=row, column=3, value=note)
            c.font = Font(name=FONT_NAME)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = 300
            row += 1
    set_col_widths(ws, {"A": 14, "B": 16, "C": 120})

    # --- Transcripts_Reference ---
    ws = wb.create_sheet("Transcripts_Reference")
    headers = ["encounter_id", "chief_complaint", "transcript", "gold_reference_note"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    row = 2
    for enc_id in SAMPLE_ENCOUNTERS:
        enc = processed[enc_id]
        ws.cell(row=row, column=1, value=enc_id).font = Font(name=FONT_NAME)
        ws.cell(row=row, column=2, value=enc.get("chief_complaint_meta") or "").font = Font(name=FONT_NAME)
        c = ws.cell(row=row, column=3, value=format_dialogue(enc["dialogue"]))
        c.font = Font(name=FONT_NAME)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c2 = ws.cell(row=row, column=4, value=enc["note_raw"])
        c2.font = Font(name=FONT_NAME)
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 400
        row += 1
    set_col_widths(ws, {"A": 14, "B": 20, "C": 90, "D": 90})

    # --- Sentence_Annotations ---
    ws = wb.create_sheet("Sentence_Annotations")
    headers = ["encounter_id", "strategy", "unit_type", "note_text", "error_type",
               "severity", "evidence_quote", "annotator_notes"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    row = 2
    first_data_row = row
    for enc_id in SAMPLE_ENCOUNTERS:
        for strategy in STRATEGIES:
            note = generated[strategy][enc_id]["generated_note"]
            for kind, text in split_note_into_units(note):
                ws.cell(row=row, column=1, value=enc_id).font = Font(name=FONT_NAME)
                ws.cell(row=row, column=2, value=strategy).font = Font(name=FONT_NAME)
                ws.cell(row=row, column=3, value=kind).font = Font(name=FONT_NAME)
                c = ws.cell(row=row, column=4, value=text)
                c.font = Font(name=FONT_NAME, bold=(kind == "header"))
                c.alignment = Alignment(wrap_text=True, vertical="top")
                if kind == "sentence":
                    for col in (5, 6, 7, 8):
                        ws.cell(row=row, column=col).fill = INPUT_FILL
                row += 1
    last_data_row = row - 1
    set_col_widths(ws, {"A": 12, "B": 14, "C": 10, "D": 70, "E": 14, "F": 12, "G": 45, "H": 35})
    for r in range(first_data_row, last_data_row + 1):
        ws.row_dimensions[r].height = 30

    dv_error = DataValidation(type="list", formula1='"none,fabrication,contradiction"', allow_blank=True)
    dv_severity = DataValidation(type="list", formula1='"minor,moderate,major"', allow_blank=True)
    ws.add_data_validation(dv_error)
    ws.add_data_validation(dv_severity)
    dv_error.add(f"E{first_data_row}:E{last_data_row}")
    dv_severity.add(f"F{first_data_row}:F{last_data_row}")

    # --- Omission_Check ---
    ws = wb.create_sheet("Omission_Check")
    headers = ["encounter_id", "strategy", "omitted_facts", "severity", "annotator_notes"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    row = 2
    om_first = row
    for enc_id in SAMPLE_ENCOUNTERS:
        for strategy in STRATEGIES:
            ws.cell(row=row, column=1, value=enc_id).font = Font(name=FONT_NAME)
            ws.cell(row=row, column=2, value=strategy).font = Font(name=FONT_NAME)
            for col in (3, 4, 5):
                ws.cell(row=row, column=col).fill = INPUT_FILL
            ws.row_dimensions[row].height = 60
            row += 1
    om_last = row - 1
    set_col_widths(ws, {"A": 12, "B": 14, "C": 70, "D": 12, "E": 35})
    dv_severity2 = DataValidation(type="list", formula1='"minor,moderate,major"', allow_blank=True)
    ws.add_data_validation(dv_severity2)
    dv_severity2.add(f"D{om_first}:D{om_last}")

    # --- Progress ---
    ws = wb.create_sheet("Progress")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 14
    rows_labels = [
        ("Metric", "Value"),
        ("Total sentences to review", f"=COUNTA(Sentence_Annotations!D{first_data_row}:D{last_data_row})"
                                       f"-COUNTIF(Sentence_Annotations!C{first_data_row}:C{last_data_row},\"header\")"),
        ("Sentences labeled (any value)", f"=COUNTA(Sentence_Annotations!E{first_data_row}:E{last_data_row})"),
        ("Labeled: fabrication", f'=COUNTIF(Sentence_Annotations!E{first_data_row}:E{last_data_row},"fabrication")'),
        ("Labeled: contradiction", f'=COUNTIF(Sentence_Annotations!E{first_data_row}:E{last_data_row},"contradiction")'),
        ("Labeled: none", f'=COUNTIF(Sentence_Annotations!E{first_data_row}:E{last_data_row},"none")'),
        ("Omission rows with notes filled in", f"=COUNTA(Omission_Check!C{om_first}:C{om_last})"),
    ]
    for i, (label, value) in enumerate(rows_labels, start=1):
        c1 = ws.cell(row=i, column=1, value=label)
        c1.font = Font(name=FONT_NAME, bold=(i == 1))
        c2 = ws.cell(row=i, column=2, value=value)
        c2.font = Font(name=FONT_NAME, bold=(i == 1))
        if i == 1:
            c1.fill = HEADER_FILL
            c2.fill = HEADER_FILL

    wb.move_sheet("Progress", offset=-5)  # order: Instructions, Progress, Notes, Transcripts, Sentence, Omission

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"wrote {OUT_PATH}")
    print(f"Sentence_Annotations data rows: {first_data_row}..{last_data_row}")
    print(f"Omission_Check data rows: {om_first}..{om_last}")


if __name__ == "__main__":
    build()
